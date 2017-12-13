#!/usr/bin/env python
# -*- coding: utf-8 -*-

from PIL import Image
import pytesseract
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import time

#dirName = 'idcard' 如果需要远程操作在配置

def get_info(image):

    image = Image.open(image)

    text = pytesseract.image_to_string(image, lang='chi_sim').replace(' ', '')
    print(text)

    name = ''.join(re.findall(r'姓名(.+?)\n', text))

    id = ''.join(re.findall(r'\d{18}|\d{17}[X|x]', text))

    return name, id

def get_uid():
    #获取唯一值
    return str(int(time.time() * 1000)) + str(int(time.clock() * 1000000))

def get_imlist(path):

    return [os.path.join(path, f) for f in os.listdir(path) if f.endswith('.jpg')]


if __name__ == '__main__':

    filename = get_imlist("idcard")

    if os.path.exists('processed.xlsx'):
        wbk = load_workbook('processed.xlsx')
        sheet = wbk.get_sheet_by_name('Nameid')
    else:
        wbk = Workbook()
        sheet = wbk.active
        sheet.title = 'Nameid'
        sheet.append(('Name', 'ID', 'UID'))
    for z in range(len(filename)):
        group = get_info(filename[z])
        #print(filename[z])
        uid = get_uid()
        print((group[0], group[1], uid))
        sheet.append((group[0], group[1], uid))
        os.rename(filename[z], '3782/'+uid+'.jpg')
    wbk.save('processed.xlsx')





